import xlrd
from openpyxl import load_workbook

workbook_object = load_workbook(filename=r'/Users/flintylemming/Downloads/test.xlsx')

sheet_object = workbook_object.worksheets[0]

print(sheet_object.cell(row=1, column=1).value)
